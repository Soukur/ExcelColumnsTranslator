"""
Excel操作を担当するモジュール
"""
from openpyxl import load_workbook, Workbook
import os
from typing import Optional
from utils import number_to_excel_column

class ExcelHandler:
    def __init__(self, input_path: str, output_path: Optional[str] = None):
        if not os.path.exists(input_path):
            raise FileNotFoundError(f"入力ファイルが見つかりません: {input_path}")

        self.input_path = input_path
        self.output_path = output_path or self._generate_output_path()

        # ワークブックを読み込む
        self.wb = load_workbook(input_path)
        self.ws = self.wb.active

    def _generate_output_path(self) -> str:
        """デフォルトの出力パスを生成"""
        dir_name = os.path.dirname(self.input_path)
        base_name = os.path.basename(self.input_path)
        name, ext = os.path.splitext(base_name)
        return os.path.join(dir_name, f"{name}_translated{ext}")

    def get_cell_value(self, row: int, col: int) -> str:
        """セルの値を取得"""
        try:
            cell = self.ws.cell(row=row, column=col)
            return str(cell.value) if cell.value is not None else ""
        except Exception as e:
            raise Exception(f"セルの読み取りに失敗しました (行: {row}, 列: {col}): {str(e)}")

    def set_cell_value(self, row: int, col: int, value: str) -> None:
        """セルに値を設定"""
        try:
            self.ws.cell(row=row, column=col, value=value)
        except Exception as e:
            raise Exception(f"セルの書き込みに失敗しました (行: {row}, 列: {col}): {str(e)}")

    def get_sheet_name(self) -> str:
        """現在のシート名を取得"""
        return self.ws.title

    def get_cell_address(self, row: int, col: int) -> str:
        """セルのアドレスを取得 (例: A1, B2)"""
        return f"{number_to_excel_column(col)}{row}"

    def save(self) -> None:
        """ワークブックを保存"""
        try:
            self.wb.save(self.output_path)
        except Exception as e:
            raise Exception(f"ファイルの保存に失敗しました: {str(e)}")